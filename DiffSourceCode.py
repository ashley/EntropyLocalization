from openpyxl import Workbook
import subprocess
import os, sys

wb = Workbook()
ws = wb.active

def saveWorkbook():
    wb.save("save.xlsx")

def prepareDirectory():
    diffPath = "/Users/ashleychen/Desktop/EntropyLocalization/Copies/Time"
    #diffPath = raw_input("Enter absolute path to aggregated source code directory: ")
    versions = [version for version in os.listdir(diffPath) if version != ".DS_Store"] #filter out default mac files
    #for versionDirectory in versions:
    versionDirectory = versions[0]
    getDiffData(diffPath + "/" + versionDirectory, versionDirectory)

def getDiffData(versionPath, versionID):
    before_files = [i for i in os.listdir(versionPath + "/b") if i != ".DS_store"]
    after_files = [i for i in os.listdir(versionPath + "/f") if i != ".DS_Store"]
    version_information = {
            "VersionID" : versionID,
            "path" : versionPath,
            "working" : str(len(before_files) == len(after_files)),
            "results" : ""
            }
    bug_information = {}
    if before_files == after_files:
        for fileName in before_files:
            diffCommand = ''.join([
                "diff ", 
                versionPath, "/b/", fileName,  " ",  
                versionPath, "/f/", fileName
                ])
            process = subprocess.Popen(diffCommand.split(), stdout=subprocess.PIPE)
            output, error = process.communicate()
            bug_information[fileName] = output
    parseOutput(bug_information["Partial.java"])
    #appendWorkbook(version_information, bug_information)

def appendWorkbook(versionMap, bugMap):
    appending = [ v for v in bugMap.values() ]
    ws.append(appending)

def parseOutput(result):
    result = result.split("\n")
    result[0]

def main():
    prepareDirectory()
    saveWorkbook()

main()
